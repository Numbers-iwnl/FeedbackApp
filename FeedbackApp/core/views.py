# -*- coding: utf-8 -*-
import csv
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncDate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator

from .forms import FeedbackForm, StatusForm
from .models import Feedback, FeedbackAttachment, FeedbackComment

# ---- Excel (openpyxl) ----
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ---- PDF (ReportLab) ----
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# ============ Access control ============
def _in_support(user):
    """Active user AND (superuser OR in group 'Suporte')."""
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name="Suporte").exists())

support_required = user_passes_test(_in_support, login_url="login")
# =======================================


# ============ Helpers ============
def _month_bounds(ym: str):
    """ym: 'YYYY-MM' → timezone-aware [start, end) for that month."""
    y, m = map(int, ym.split("-"))
    start_naive = datetime(y, m, 1)
    end_naive = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
    start = timezone.make_aware(start_naive)
    end = timezone.make_aware(end_naive)
    return start, end


def _current_operator(user):
    """Pretty operator name."""
    full = (user.get_full_name() or "").strip()
    if full:
        return full
    if (user.first_name or "").strip():
        return user.first_name.strip()
    return user.username


def _filtered_queryset(request):
    """
    Apply list filters from querystring to Feedback queryset.
    Supports: aluno, operador, curso, tipo, assunto, status, de (YYYY-MM-DD), ate (YYYY-MM-DD).
    """
    qs = Feedback.objects.all()
    g = request.GET

    aluno = (g.get("aluno") or "").strip()
    operador = (g.get("operador") or "").strip()
    curso = (g.get("curso") or "").strip()
    tipo = (g.get("tipo") or "").strip()
    assunto = (g.get("assunto") or "").strip()
    status = (g.get("status") or "").strip()
    de = (g.get("de") or "").strip()
    ate = (g.get("ate") or "").strip()

    if aluno:
        qs = qs.filter(student_name__icontains=aluno)
    if operador:
        qs = qs.filter(operator_name__icontains=operador)
    if curso:
        qs = qs.filter(course_name__icontains=curso)
    if tipo:
        qs = qs.filter(type=tipo)
    if assunto:
        qs = qs.filter(subject=assunto)
    if status:
        qs = qs.filter(status=status)

    fmt = "%Y-%m-%d"
    if de:
        try:
            start_naive = datetime.strptime(de, fmt)
            start = timezone.make_aware(start_naive)
            qs = qs.filter(created_at__gte=start)
        except ValueError:
            pass
    if ate:
        try:
            d = datetime.strptime(ate, fmt)
            end = timezone.make_aware(datetime(d.year, d.month, d.day, 23, 59, 59, 999999))
            qs = qs.filter(created_at__lte=end)
        except ValueError:
            pass

    return qs
# =======================================


@login_required
@support_required
def ping(request):
    return HttpResponse("core ok")


# ---- Create ----
@login_required
@support_required
def feedback_create(request):
    if request.method == "POST":
        form = FeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            fb = form.save(commit=False)
            fb.operator_name = _current_operator(request.user)  # force actual operator
            fb.save()

            messages.success(request, f"Feedback #{fb.id} criado com sucesso.")

            for f in form.cleaned_data.get("attachments", []):
                FeedbackAttachment.objects.create(
                    feedback=fb,
                    file=f,
                    mime_type=getattr(f, "content_type", None),
                    file_size=getattr(f, "size", None),
                )
            return redirect(f"{reverse('feedback_create')}?created_id={fb.id}")
        return render(request, "core/feedback_form.html", {"form": form})

    # GET — prefill operator
    initial = {"operator_name": _current_operator(request.user)}
    return render(
        request,
        "core/feedback_form.html",
        {
            "form": FeedbackForm(initial=initial),
            "created_id": request.GET.get("created_id"),
        },
    )


# ---- List (filters + pagination) ----
@login_required
@support_required
def feedback_list(request):
    qs = _filtered_queryset(request).order_by("-created_at")

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    q = request.GET.copy()
    q.pop("page", None)
    querystring = q.urlencode()

    context = {
        "items": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "querystring": querystring,

        "aluno": (request.GET.get("aluno") or "").strip(),
        "operador": (request.GET.get("operador") or "").strip(),
        "curso": (request.GET.get("curso") or "").strip(),
        "tipo": (request.GET.get("tipo") or "").strip(),
        "assunto": (request.GET.get("assunto") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
        "de": (request.GET.get("de") or "").strip(),
        "ate": (request.GET.get("ate") or "").strip(),

        "tipo_choices": Feedback.TIPO_CHOICES,
        "assunto_choices": Feedback.ASSUNTO_CHOICES,
        "status_choices": Feedback.STATUS_CHOICES,
    }
    return render(request, "core/feedback_list.html", context)


# ---- Detail (status + comments) ----
@login_required
@support_required
def feedback_detail(request, pk):
    fb = get_object_or_404(Feedback, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "status":
            form_status = StatusForm(request.POST)
            if form_status.is_valid():
                new_status = form_status.cleaned_data["status"]
                fb.status = new_status
                fb.resolved_at = fb.resolved_at or timezone.now() if new_status == "resolvido" else None
                fb.save()
                messages.success(request, f"Status atualizado para {fb.get_status_display()}.")
            return redirect("feedback_detail", pk=fb.pk)

        if action == "comment":
            author = (request.POST.get("author_name") or "Suporte").strip()
            text = (request.POST.get("comment_text") or "").strip()
            if text:
                FeedbackComment.objects.create(
                    feedback=fb,
                    author_name=author,
                    comment_text=text,
                )
                messages.success(request, "Comentário adicionado.")
            return redirect("feedback_detail", pk=fb.pk)

    form_status = StatusForm(initial={"status": fb.status})
    return render(request, "core/feedback_detail.html", {"fb": fb, "form_status": form_status})


# ---- Attachment gated download ----
@login_required
@support_required
def attachment_download(request, pk):
    att = get_object_or_404(FeedbackAttachment, pk=pk)
    try:
        return FileResponse(
            open(att.file.path, "rb"),
            as_attachment=False,
            content_type=att.mime_type or "application/octet-stream",
            filename=Path(att.file.name).name,
        )
    except FileNotFoundError:
        raise Http404("Arquivo não encontrado")


# ---- CSV export (respects filters) ----
@login_required
@support_required
def export_csv(request):
    qs = _filtered_queryset(request).order_by("-created_at")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = "attachment; filename=feedbacks.csv"
    w = csv.writer(response, delimiter=";")

    w.writerow([
        "id", "data", "aluno", "operador", "tipo", "assunto",
        "curso", "turma", "status", "descricao", "anexos_qtd"
    ])

    for f in qs:
        w.writerow([
            f.id,
            f.created_at.strftime("%Y-%m-%d %H:%M"),
            f.student_name or "",
            f.operator_name or "",
            f.type,
            f.subject,
            f.course_name or "",
            f.class_name or "",
            f.status,
            (f.description or "").replace("\n", " ").strip()[:500],
            f.attachments.count(),
        ])

    return response


# ---- Excel export (respects filters) ----
@login_required
@support_required
def export_excel(request):
    qs = _filtered_queryset(request).order_by("-created_at")[:2000]  # safety cap

    wb = Workbook()
    ws = wb.active
    ws.title = "Feedbacks"

    headers = [
        "ID", "Data", "Aluno", "Operador", "Tipo", "Assunto",
        "Curso", "Turma", "Status", "Descrição", "Anexos (qtd)"
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    tipo_map = dict(Feedback.TIPO_CHOICES)
    assunto_map = dict(Feedback.ASSUNTO_CHOICES)
    status_map = dict(Feedback.STATUS_CHOICES)

    for f in qs:
        ws.append([
            f.id,
            f.created_at.strftime("%Y-%m-%d %H:%M"),
            f.student_name or "",
            f.operator_name or "",
            tipo_map.get(f.type, f.type),
            assunto_map.get(f.subject, f.subject),
            f.course_name or "",
            f.class_name or "",
            status_map.get(f.status, f.status),
            (f.description or "").replace("\n", " ").strip(),
            f.attachments.count(),
        ])

    # auto width
    for col_idx in range(1, len(headers) + 1):
        col = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col]:
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = "attachment; filename=feedbacks.xlsx"
    return resp


# ---- PDF export (monthly summary) ----
@login_required
@support_required
def export_pdf(request):
    month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    start, end = _month_bounds(month)
    q_month = Feedback.objects.filter(created_at__gte=start, created_at__lt=end)
    qs = Feedback.objects.order_by("-created_at")[:200]

    resumo = {
        "Total": q_month.count(),
        "Elogios": q_month.filter(type="elogio").count(),
        "Reclamações": q_month.filter(type="reclamacao").count(),
        "Sugestões": q_month.filter(type="sugestao").count(),
        "Resolvidos": q_month.filter(status="resolvido").count(),
    }

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Relatório de Feedbacks", styles["Title"]))
    story.append(Paragraph(f"Mês: {month}", styles["Normal"]))
    story.append(Spacer(1, 10))

    data_resumo = [["Métrica", "Valor"]] + [[k, v] for k, v in resumo.items()]
    t1 = Table(data_resumo, colWidths=[220, 120])
    t1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    story.append(t1)
    story.append(Spacer(1, 12))

    headers = ["ID", "Data", "Aluno", "Tipo", "Assunto", "Curso", "Turma", "Status"]
    rows = []
    tipo_map = dict(Feedback.TIPO_CHOICES)
    assunto_map = dict(Feedback.ASSUNTO_CHOICES)
    status_map = dict(Feedback.STATUS_CHOICES)

    for f in qs:
        rows.append([
            f.id,
            f.created_at.strftime("%Y-%m-%d %H:%M"),
            f.student_name or "",
            tipo_map.get(f.type, f.type),
            assunto_map.get(f.subject, f.subject),
            f.course_name or "",
            f.class_name or "",
            status_map.get(f.status, f.status),
        ])

    t2 = Table([headers] + rows, repeatRows=1)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(Paragraph("Detalhes (últimos 200)", styles["Heading3"]))
    story.append(t2)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="feedbacks.pdf"'
    return resp


# ---- Monthly stats (JSON) ----
@login_required
@support_required
def stats_summary(request):
    month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    start, end = _month_bounds(month)
    qs = Feedback.objects.filter(created_at__gte=start, created_at__lt=end)

    # average resolution time (hours) for resolved
    qs_res = qs.filter(status="resolvido", resolved_at__isnull=False)
    if qs_res.exists():
        qs_res = qs_res.annotate(
            dur=ExpressionWrapper(F("resolved_at") - F("created_at"), output_field=DurationField())
        )
        avg_dur = qs_res.aggregate(a=Avg("dur"))["a"]
        avg_hours = round(avg_dur.total_seconds() / 3600, 1) if avg_dur else 0.0
    else:
        avg_hours = 0.0

    data = {
        "elogios": qs.filter(type="elogio").count(),
        "reclamacoes": qs.filter(type="reclamacao").count(),
        "sugestoes": qs.filter(type="sugestao").count(),
        "resolvidos": qs.filter(status="resolvido").count(),
        "total": qs.count(),
        "avg_resolution_hours": avg_hours,
    }
    data["taxa_resolucao_pct"] = round(100 * data["resolvidos"] / data["total"], 1) if data["total"] else 0.0
    return JsonResponse(data)


@login_required
@support_required
def stats_breakdown(request):
    month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    start, end = _month_bounds(month)
    qs = Feedback.objects.filter(created_at__gte=start, created_at__lt=end)

    by_type     = list(qs.values("type").annotate(n=Count("id")))
    by_subject  = list(qs.values("subject").annotate(n=Count("id")))
    by_course   = list(qs.values("course_name").annotate(n=Count("id")))
    by_status   = list(qs.values("status").annotate(n=Count("id")))
    by_operator = list(qs.values("operator_name").annotate(n=Count("id")).order_by("-n")[:8])

    label_type    = dict(Feedback.TIPO_CHOICES)
    label_subject = dict(Feedback.ASSUNTO_CHOICES)
    label_status  = dict(Feedback.STATUS_CHOICES)

    def map_labels(rows, label_map, field):
        out = []
        for r in rows:
            key = r[field]
            label = label_map.get(key, key if key else "—")
            out.append({"label": label, "value": r["n"]})
        return out

    return JsonResponse({
        "type":     map_labels(by_type,    label_type,    "type"),
        "subject":  map_labels(by_subject, label_subject, "subject"),
        "status":   map_labels(by_status,  label_status,  "status"),
        "course":   [{"label": r["course_name"] or "—", "value": r["n"]} for r in by_course],
        "operator": [{"label": r["operator_name"] or "—", "value": r["n"]} for r in by_operator],
    })


@login_required
@support_required
def stats_timeseries(request):
    """Daily totals vs resolved for selected month."""
    month = request.GET.get("month") or timezone.now().strftime("%Y-%m")
    start, end = _month_bounds(month)
    qs = Feedback.objects.filter(created_at__gte=start, created_at__lt=end)

    total_by_day = qs.annotate(d=TruncDate("created_at")).values("d").annotate(n=Count("id"))
    resolved_by_day = qs.filter(status="resolvido").annotate(d=TruncDate("created_at")).values("d").annotate(n=Count("id"))

    map_total = {row["d"].isoformat(): row["n"] for row in total_by_day}
    map_res   = {row["d"].isoformat(): row["n"] for row in resolved_by_day}

    labels, total, resolved, pending = [], [], [], []
    cur = start.date()
    while cur < end.date():
        s = cur.isoformat()
        t = map_total.get(s, 0)
        r = map_res.get(s, 0)
        p = max(t - r, 0)
        labels.append(s)
        total.append(t)
        resolved.append(r)
        pending.append(p)
        cur = cur + timedelta(days=1)

    return JsonResponse({"labels": labels, "total": total, "resolved": resolved, "pending": pending})


@login_required
@support_required
def stats_recent(request):
    """Last few feedbacks (compact for dashboard)."""
    qs = Feedback.objects.order_by("-created_at")[:8]
    data = [
        {
            "id": f.id,
            "student": f.student_name,
            "status": f.get_status_display(),
            "status_key": f.status,
            "subject": f.get_subject_display(),
            "created": f.created_at.strftime("%Y-%m-%d %H:%M"),
        }
        for f in qs
    ]
    return JsonResponse({"items": data})


@login_required
@support_required
def dashboard(request):
    return render(request, "core/dashboard.html")
